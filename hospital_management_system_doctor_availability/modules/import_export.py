"""Upload / Download module for Hospital Management System.

Features:
- Upload Patients Data from CSV, JSON, Excel (.xlsx), or TXT files.
- Download Patients Data to JSON, CSV, Excel (.xlsx), or TXT using a Save As window.
- Upload Doctors Data from CSV, JSON, Excel (.xlsx), or TXT files.
- Download Doctors Data to JSON, CSV, Excel (.xlsx), or TXT using a Save As window.

Note:
This module uses tkinter file dialogs, so it should be run on a system with GUI support
like Windows / VS Code terminal. It will not open a file window in some server-only terminals.
"""

import csv
import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from utils.id_generator import get_next_id
from utils.input_helper import print_line, pause


SUPPORTED_UPLOAD_TYPES = (
    ("Supported data files", "*.csv *.json *.xlsx *.txt"),
    ("CSV files", "*.csv"),
    ("JSON files", "*.json"),
    ("Excel files", "*.xlsx"),
    ("Text files", "*.txt"),
    ("All files", "*.*"),
)

SUPPORTED_DOWNLOAD_TYPES = (
    ("JSON files", "*.json"),
    ("CSV files", "*.csv"),
    ("Excel files", "*.xlsx"),
    ("Text files", "*.txt"),
    ("All files", "*.*"),
)

SUPPORTED_DOWNLOAD_EXTENSIONS = {".json", ".csv", ".xlsx", ".txt"}


# -----------------------------------------------------------------------------
# File dialog helpers
# -----------------------------------------------------------------------------
def _open_file_dialog(title: str) -> Optional[Path]:
    """Open Windows file picker and return selected file path."""
    try:
        from tkinter import Tk, filedialog

        root = Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        file_path = filedialog.askopenfilename(title=title, filetypes=SUPPORTED_UPLOAD_TYPES)
        root.destroy()
        return Path(file_path) if file_path else None
    except Exception as exc:
        print("Unable to open file selection window.")
        print("Reason:", exc)
        return None


def _save_file_dialog(title: str, default_file_name: str) -> Optional[Path]:
    """Open Save As window and return selected save path.

    User can save as .json, .csv, .xlsx, or .txt.
    The selected/typed extension decides the download format.
    """
    try:
        from tkinter import Tk, filedialog

        root = Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        file_path = filedialog.asksaveasfilename(
            title=title,
            initialfile=default_file_name,
            defaultextension=".json",
            filetypes=SUPPORTED_DOWNLOAD_TYPES,
        )
        root.destroy()
        if not file_path:
            return None

        path = Path(file_path)
        if path.suffix.lower() not in SUPPORTED_DOWNLOAD_EXTENSIONS:
            path = path.with_suffix(".json")
        return path
    except Exception as exc:
        print("Unable to open save location window.")
        print("Reason:", exc)
        return None


# -----------------------------------------------------------------------------
# Data reading helpers
# -----------------------------------------------------------------------------
def _normalize_key(key: Any) -> str:
    """Convert file header names into database-style keys."""
    return str(key or "").strip().upper().replace(" ", "_").replace("-", "_")


def _normalize_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Normalize all keys in one row."""
    return {_normalize_key(k): v for k, v in row.items()}


def _clean(value: Any) -> Optional[Any]:
    """Convert blank values to None and trim strings."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def _to_int(value: Any) -> Optional[int]:
    value = _clean(value)
    if value is None:
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def _to_float(value: Any) -> Optional[float]:
    value = _clean(value)
    if value is None:
        return None
    try:
        return float(value)
    except Exception:
        return None


def _get(row: Dict[str, Any], *keys: str) -> Optional[Any]:
    """Get first matching key value from normalized row."""
    for key in keys:
        value = row.get(_normalize_key(key))
        if _clean(value) is not None:
            return _clean(value)
    return None


def _read_csv_file(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        return [_normalize_row(row) for row in reader]


def _read_json_file(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig") as file:
        data = json.load(file)

    # Accept direct list: [{...}, {...}]
    if isinstance(data, list):
        return [_normalize_row(row) for row in data if isinstance(row, dict)]

    # Accept dictionary formats like:
    # {"patients": [...]}, {"doctors": [...]}, {"data": [...]}, {"records": [...]}
    if isinstance(data, dict):
        for key in ("patients", "doctors", "data", "records", "items"):
            value = data.get(key)
            if isinstance(value, list):
                return [_normalize_row(row) for row in value if isinstance(row, dict)]

        # Accept single record: {"PATIENT_NAME": "Ravi", ...}
        return [_normalize_row(data)]

    return []


def _read_excel_file(path: Path) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Excel upload requires openpyxl. Install it using: pip install openpyxl")
        return []

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        return []

    header_index = None
    headers = []
    for index, row in enumerate(rows):
        if row and any(cell is not None and str(cell).strip() for cell in row):
            header_index = index
            headers = [_normalize_key(cell) for cell in row]
            break

    if header_index is None:
        return []

    records = []
    for row in rows[header_index + 1:]:
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        record = {}
        for index, header in enumerate(headers):
            if header:
                record[header] = row[index] if index < len(row) else None
        records.append(record)
    return records


def _read_text_file(path: Path) -> List[Dict[str, Any]]:
    """Read TXT file as delimited text. Supports comma, tab, pipe, and semicolon."""
    text = path.read_text(encoding="utf-8-sig").strip()
    if not text:
        return []

    # If TXT contains JSON content, read it as JSON.
    if text.startswith("[") or text.startswith("{"):
        return _read_json_file(path)

    lines = text.splitlines()
    sample = "\n".join(lines[:5])

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t|;")
    except Exception:
        dialect = csv.excel
        dialect.delimiter = ","

    reader = csv.DictReader(lines, dialect=dialect)
    return [_normalize_row(row) for row in reader]


def _load_records_from_file(path: Path) -> List[Dict[str, Any]]:
    """Detect file type and load rows."""
    extension = path.suffix.lower()

    if extension == ".csv":
        print("File type detected: CSV")
        return _read_csv_file(path)
    if extension == ".json":
        print("File type detected: JSON")
        return _read_json_file(path)
    if extension == ".xlsx":
        print("File type detected: Excel")
        return _read_excel_file(path)
    if extension == ".txt":
        print("File type detected: TXT")
        return _read_text_file(path)

    print("Unsupported file type.")
    print("Please select CSV, JSON, Excel .xlsx, or TXT file.")
    return []


# -----------------------------------------------------------------------------
# Download writing helpers
# -----------------------------------------------------------------------------
def _format_value(value: Any) -> str:
    """Convert database values into clean file output values."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _json_safe(value: Any) -> Any:
    """Make Oracle/date/decimal values safe for JSON download."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, Decimal):
        return float(value)
    return value


def _write_json(path: Path, data: List[Dict[str, Any]]) -> None:
    safe_data = [{key: _json_safe(value) for key, value in row.items()} for row in data]
    with path.open("w", encoding="utf-8") as file:
        json.dump(safe_data, file, indent=4)


def _write_csv(path: Path, data: List[Dict[str, Any]]) -> None:
    if not data:
        path.write_text("", encoding="utf-8")
        return

    columns = list(data[0].keys())
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=columns)
        writer.writeheader()
        for row in data:
            writer.writerow({column: _format_value(row.get(column)) for column in columns})


def _write_excel(path: Path, data: List[Dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError("Excel download requires openpyxl. Install it using: pip install openpyxl")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"

    if data:
        columns = list(data[0].keys())
        sheet.append([column.upper() for column in columns])
        for row in data:
            sheet.append([_format_value(row.get(column)) for column in columns])
    else:
        sheet.append(["NO_RECORDS_FOUND"])

    workbook.save(path)


def _write_txt(path: Path, data: List[Dict[str, Any]]) -> None:
    if not data:
        path.write_text("No records found.\n", encoding="utf-8")
        return

    columns = list(data[0].keys())
    with path.open("w", encoding="utf-8") as file:
        file.write(" | ".join(column.upper() for column in columns) + "\n")
        file.write("=" * 120 + "\n")
        for row in data:
            file.write(" | ".join(_format_value(row.get(column)) for column in columns) + "\n")


def _download_data_file(path: Path, data: List[Dict[str, Any]], label: str) -> None:
    """Write download data based on selected file extension."""
    path.parent.mkdir(parents=True, exist_ok=True)
    extension = path.suffix.lower()

    if extension == ".json":
        _write_json(path, data)
    elif extension == ".csv":
        _write_csv(path, data)
    elif extension == ".xlsx":
        _write_excel(path, data)
    elif extension == ".txt":
        _write_txt(path, data)
    else:
        raise ValueError("Unsupported download type. Use .json, .csv, .xlsx, or .txt")

    print(f"{label} data downloaded successfully as {extension.upper().replace('.', '')} file:")
    print(path)


# -----------------------------------------------------------------------------
# Upload Patients
# -----------------------------------------------------------------------------
def upload_patients_data(conn):
    cursor = conn.cursor()

    print("\nUPLOAD PATIENTS DATA")
    print_line()
    print("Supported file types: CSV, JSON, Excel (.xlsx), TXT")
    print("Opening file selection window...")

    path = _open_file_dialog("Select Patient Data File")
    if not path:
        print("Upload cancelled. No file selected.")
        return

    if not path.exists():
        print("Selected file was not found.")
        return

    print(f"Selected File: {path}")
    records = _load_records_from_file(path)

    if not records:
        print("No valid records found in the selected file.")
        return

    inserted = 0
    failed = 0

    print("\nUploading patients data...")

    try:
        for row in records:
            try:
                patient_name = _get(row, "PATIENT_NAME", "NAME")
                phone = _get(row, "PHONE", "MOBILE", "MOBILE_NUMBER", "CONTACT")

                if not patient_name or not phone:
                    raise ValueError("PATIENT_NAME and PHONE are required for patient upload.")

                patient_id = get_next_id(cursor, "PATIENT_MASTER", "PATIENT_ID", "P", 3)

                cursor.execute(
                    """
                    INSERT INTO PATIENT_MASTER (
                        PATIENT_ID, PATIENT_NAME, AGE, GENDER, PHONE, EMAIL, ADDRESS,
                        CITY, BLOOD_GROUP, EMERGENCY_CONTACT, REGISTRATION_DATE
                    ) VALUES (
                        :patient_id, :patient_name, :age, :gender, :phone, :email, :address,
                        :city, :blood_group, :emergency_contact, SYSDATE
                    )
                    """,
                    {
                        "patient_id": patient_id,
                        "patient_name": patient_name,
                        "age": _to_int(_get(row, "AGE")),
                        "gender": _get(row, "GENDER", "SEX"),
                        "phone": phone,
                        "email": _get(row, "EMAIL", "EMAIL_ID"),
                        "address": _get(row, "ADDRESS"),
                        "city": _get(row, "CITY"),
                        "blood_group": _get(row, "BLOOD_GROUP", "BLOOD"),
                        "emergency_contact": _get(row, "EMERGENCY_CONTACT", "EMERGENCY_PHONE"),
                    },
                )
                inserted += 1
            except Exception as exc:
                failed += 1
                print("Failed patient row:", row, "Error:", exc)

        conn.commit()
        print("\nUpload completed.")
        print(f"Inserted: {inserted}")
        print(f"Failed  : {failed}")

    except Exception as exc:
        conn.rollback()
        print("Patient upload failed:", exc)


# -----------------------------------------------------------------------------
# Download Patients
# -----------------------------------------------------------------------------
def download_patients_data(conn):
    cursor = conn.cursor()

    print("\nDOWNLOAD PATIENTS DATA")
    print_line()
    print("Choose the file type in the Save As window: JSON, CSV, Excel, or TXT")
    print("Opening save location window...")

    path = _save_file_dialog(
        title="Save Patients Data",
        default_file_name="patients_data.json",
    )

    if not path:
        print("Download cancelled. No save location selected.")
        return

    cursor.execute(
        """
        SELECT PATIENT_ID, PATIENT_NAME, AGE, GENDER, PHONE, EMAIL, ADDRESS, CITY,
               BLOOD_GROUP, EMERGENCY_CONTACT, REGISTRATION_DATE
        FROM PATIENT_MASTER
        ORDER BY PATIENT_ID
        """
    )

    columns = [description[0].lower() for description in cursor.description]
    data = [dict(zip(columns, row)) for row in cursor.fetchall()]

    try:
        _download_data_file(path, data, "Patients")
    except Exception as exc:
        print("Patients download failed:", exc)


# -----------------------------------------------------------------------------
# Upload Doctors
# -----------------------------------------------------------------------------
def upload_doctors_data(conn):
    cursor = conn.cursor()

    print("\nUPLOAD DOCTORS DATA")
    print_line()
    print("Supported file types: CSV, JSON, Excel (.xlsx), TXT")
    print("Opening file selection window...")

    path = _open_file_dialog("Select Doctor Data File")
    if not path:
        print("Upload cancelled. No file selected.")
        return

    if not path.exists():
        print("Selected file was not found.")
        return

    print(f"Selected File: {path}")
    records = _load_records_from_file(path)

    if not records:
        print("No valid records found in the selected file.")
        return

    inserted = 0
    failed = 0

    print("\nUploading doctors data...")

    try:
        for row in records:
            try:
                doctor_name = _get(row, "DOCTOR_NAME", "NAME")
                specialization = _get(row, "SPECIALIZATION", "DEPARTMENT")
                phone = _get(row, "PHONE", "MOBILE", "MOBILE_NUMBER", "CONTACT")

                if not doctor_name or not specialization or not phone:
                    raise ValueError("DOCTOR_NAME, SPECIALIZATION, and PHONE are required for doctor upload.")

                doctor_id = get_next_id(cursor, "DOCTOR_MASTER", "DOCTOR_ID", "D", 3)

                cursor.execute(
                    """
                    INSERT INTO DOCTOR_MASTER (
                        DOCTOR_ID, DOCTOR_NAME, SPECIALIZATION, PHONE, EMAIL,
                        EXPERIENCE_YEARS, CONSULTATION_FEE, AVAILABLE_DAYS,
                        AVAILABLE_FROM, AVAILABLE_TO, MAX_PATIENTS_PER_SLOT, STATUS
                    ) VALUES (
                        :doctor_id, :doctor_name, :specialization, :phone, :email,
                        :experience_years, :consultation_fee, :available_days,
                        :available_from, :available_to, :max_patients_per_slot, :status
                    )
                    """,
                    {
                        "doctor_id": doctor_id,
                        "doctor_name": doctor_name,
                        "specialization": specialization,
                        "phone": phone,
                        "email": _get(row, "EMAIL", "EMAIL_ID"),
                        "experience_years": _to_int(_get(row, "EXPERIENCE_YEARS", "EXPERIENCE")),
                        "consultation_fee": _to_float(_get(row, "CONSULTATION_FEE", "FEE")),
                        "available_days": _get(row, "AVAILABLE_DAYS", "DAYS"),
                        "available_from": _get(row, "AVAILABLE_FROM", "FROM_TIME", "START_TIME"),
                        "available_to": _get(row, "AVAILABLE_TO", "TO_TIME", "END_TIME"),
                        "max_patients_per_slot": _to_int(_get(row, "MAX_PATIENTS_PER_SLOT", "MAX_PATIENTS")),
                        "status": _get(row, "STATUS") or "Active",
                    },
                )
                inserted += 1
            except Exception as exc:
                failed += 1
                print("Failed doctor row:", row, "Error:", exc)

        conn.commit()
        print("\nUpload completed.")
        print(f"Inserted: {inserted}")
        print(f"Failed  : {failed}")

    except Exception as exc:
        conn.rollback()
        print("Doctor upload failed:", exc)


# -----------------------------------------------------------------------------
# Download Doctors
# -----------------------------------------------------------------------------
def download_doctors_data(conn):
    cursor = conn.cursor()

    print("\nDOWNLOAD DOCTORS DATA")
    print_line()
    print("Choose the file type in the Save As window: JSON, CSV, Excel, or TXT")
    print("Opening save location window...")

    path = _save_file_dialog(
        title="Save Doctors Data",
        default_file_name="doctors_data.json",
    )

    if not path:
        print("Download cancelled. No save location selected.")
        return

    cursor.execute(
        """
        SELECT DOCTOR_ID, DOCTOR_NAME, SPECIALIZATION, PHONE, EMAIL,
               EXPERIENCE_YEARS, CONSULTATION_FEE, AVAILABLE_DAYS,
               AVAILABLE_FROM, AVAILABLE_TO, MAX_PATIENTS_PER_SLOT, STATUS
        FROM DOCTOR_MASTER
        ORDER BY DOCTOR_ID
        """
    )

    columns = [description[0].lower() for description in cursor.description]
    data = [dict(zip(columns, row)) for row in cursor.fetchall()]

    try:
        _download_data_file(path, data, "Doctors")
    except Exception as exc:
        print("Doctors download failed:", exc)


# -----------------------------------------------------------------------------
# Menu
# -----------------------------------------------------------------------------
def import_export_menu(conn):
    while True:
        print("\nIMPORT / DOWNLOAD")
        print_line()
        print("1. Upload Patients Data")
        print("2. Download Patients Data")
        print("3. Upload Doctors Data")
        print("4. Download Doctors Data")
        print("5. Back to Main Menu")

        choice = input("Enter choice: ").strip()

        if choice == "1":
            upload_patients_data(conn)
            pause()
        elif choice == "2":
            download_patients_data(conn)
            pause()
        elif choice == "3":
            upload_doctors_data(conn)
            pause()
        elif choice == "4":
            download_doctors_data(conn)
            pause()
        elif choice == "5":
            break
        else:
            print("Invalid choice. Please enter a number from 1 to 5.")
